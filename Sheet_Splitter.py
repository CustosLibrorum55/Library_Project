import openpyxl
from pprint import pprint
import csv
import json

miscFile = open("Spreadsheets/miscFile.json", "w")
#miscWriter = csv.writer(miscFile)

realFile = open("Spreadsheets/realFile.json", "w")
#realWriter = csv.writer(realFile)

template = {
    "Author": "",
    "Title": "",
    "Subtitle": "",
    "Publisher": "",
    "Date": "",
    "Collection": "",
    "Edition": "",
}



def fill_template(row, template, headers):
    
    for index, cell in enumerate(row):
        if index < len(headers):
        #print(cell.internal_value)
            template[headers[index]] = cell.value
        else:
            print("EEEOR")
    return template
        #try:
        #    #header = activeSheet.cell(index + 1, index + 1).internal_value.strip()
        #    #template[] = cell.internal_value
        #    return template
        #except AttributeError as ae:
        #    print("Error")
        #    print(ae)
    
    #for cell in enumerate(row):
    #    try:
    #        header = activeSheet.cell(1, cell[0]).internal_value.strip()
    #        template[header] = cell.internal_value
    #        return template
    #    except AttributeError as e:
    #        print("Error")
    #        print(e)

collection = {}
mainSheet = openpyxl.open("Book1.xlsx", read_only=True)
#headers = [cell.value for cell in mainSheet.active.cell(row=1, column=1)]

for sheet in mainSheet.worksheets:
    miscBooks = []
    realBooks = []
    headers = []
    for cell in sheet[1]:
        try:
            headers.append(cell.value.strip())
        except AttributeError as ae:
            print(cell.value)
        
    print(headers) 
    for row in sheet.rows:
        authorName = row[0].internal_value.lower()
        if "misc" in authorName:
            filledTemplate = fill_template(row, template, headers)
            miscBooks.append(filledTemplate)
        else:
            filledTemplate = fill_template(row, template, headers)
            realBooks.append(filledTemplate)

        collection[sheet.title] = {
        "real": realBooks,
        "misc": miscBooks
        }
with open("text.json", "w") as t:
    json.dump(collection, t, indent=2) 